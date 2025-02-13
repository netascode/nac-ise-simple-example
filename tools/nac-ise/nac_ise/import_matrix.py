import openpyxl
import yaml

MATRIX_SHEET = 'Matrix'
SGACL_SHEET = 'CUSTOM SGACL'
ALLOW_COLOR = 'FF00B050'
DENY_COLOR = 'FFFF0000'
CUSTOM_COLOR = 'FF0070C0'
DEFAULT_SGTS = ['Auditors', 'BYOD', 'Contractors', 'Developers', 'Development_Servers', 'Employees', 'Guests', 'Network_Services', 'PCI_Servers', 'Point_of_Sale_Systems', 'Production_Servers', 'Production_Users', 'Quarantined_Systems', 'Test_Servers', 'TrustSec_Devices', 'Unknown']
SGT_START_VALUE = 1000

class Matrix:
    def __init__(self):
        self.matrix = []
        self.sgacls = []
        self.security_groups = []

    @classmethod
    def import_matrix(cls, input_file, output_file, sgt_start_value):
        instance = cls()
        instance.populate_matrix_dict(input_file, sgt_start_value)
        instance.dump_to_yml(output_file)
        
    def populate_matrix_dict(self, input_file, sgt_start_value):
        workbook = openpyxl.load_workbook(input_file, data_only=True)
        
        sgacl_sheet = workbook[SGACL_SHEET]
        sgacl_dict = {}
        sg = []
        for row in sgacl_sheet.iter_rows(min_row=2):
            if row[1].value is not None and row[2].value is not None:
                sgacl_dict[row[0].value] = [row[1].value,row[2].value,row[3].value]
                self.sgacls.append({
                    'name' : row[1].value,
                    'ip_version': "IPV4",
                    'description': row[3].value,
                    'acl_content' : row[2].value
                })

        sheet = workbook[MATRIX_SHEET]

        dst_sgts = [cell[0].value for cell in sheet.iter_cols(min_col=3) if cell[0].value]
        sg = dst_sgts

        for row in sheet.iter_rows(min_row=3):
            src_sgt = row[0].value
            sg.append(src_sgt)
            for i, cell in enumerate(row[2:]):
                if cell.fill.start_color.rgb == ALLOW_COLOR:
                    sgacl_name = 'Permit IP'
                elif cell.fill.start_color.rgb == DENY_COLOR:
                    sgacl_name = 'Deny IP'
                elif cell.value is not None and cell.fill.start_color.rgb == CUSTOM_COLOR:
                    sgacl_name = sgacl_dict[cell.value][0]
                else:
                     continue

                self.matrix.append({
                    'source_sgt' : src_sgt,
                    'destination_sgt' : dst_sgts[i],
                    'rule_status': 'ENABLED',
                    'sgacl_name' : sgacl_name
                })
        sg = list(set(sg))
        
        sg_value = sgt_start_value
        for i in sg:
            sg_value = sg_value + 1
            if i not in DEFAULT_SGTS and i is not None:
                self.security_groups.append({
                    'name' : i,
                    'description' : i,
                    'propagate_to_apic': False,
                    'value' : sg_value
                })

    def dump_to_yml(self, output_file):
        with open(output_file, 'w') as file:
            file.write('---\n')
            yaml.dump({
                "ise" : {
                    "trust_sec" : {
                        "matrix_entries": self.matrix,
                        "security_group_acls" : self.sgacls,
                        "security_groups" : self.security_groups
                    }
                }
            }, file, default_flow_style=False, sort_keys=False)